from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from semantic_reasoning_agent.documents.errors import DocumentProcessingError
from semantic_reasoning_agent.documents.models import DocumentIngestionOptions, ParsedChunk, ParsedDocument


PARSER_NAME = "openpyxl"
PARSER_VERSION = "xlsx-structured-v2"


class XlsxParser:
    supported_types = ("xlsx",)

    def parse(
        self,
        filename: str,
        content: bytes,
        title: str | None = None,
        *,
        options: DocumentIngestionOptions | None = None,
    ) -> ParsedDocument:
        del options
        workbook = load_workbook(BytesIO(content), data_only=True)
        chunks: list[ParsedChunk] = []
        sheet_names: list[str] = []

        for sheet in workbook.worksheets:
            sheet_names.append(sheet.title)
            non_empty_rows: list[tuple[int, list[str]]] = []
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [_stringify_cell(cell) for cell in row]
                if any(values):
                    non_empty_rows.append((row_number, values))

            if not non_empty_rows:
                continue

            header_row_number, header_values = non_empty_rows[0]
            headers = [
                header if header else f"column_{column_index}"
                for column_index, header in enumerate(header_values, start=1)
            ]
            chunks.append(
                ParsedChunk(
                    text=f"Sheet: {sheet.title}\nColumns: {', '.join(headers)}",
                    chunk_index=len(chunks),
                    section_title=sheet.title,
                    sheet_name=sheet.title,
                    row_start=header_row_number,
                    row_end=header_row_number,
                    column_headers=tuple(headers),
                    detected_table_id=f"{sheet.title}:schema",
                )
            )

            data_rows = non_empty_rows[1:]
            for offset in range(0, len(data_rows), 5):
                group = data_rows[offset : offset + 5]
                lines: list[str] = []
                for row_number, values in group:
                    pairs = [
                        f"{header}: {value}"
                        for header, value in zip(headers, values)
                        if value
                    ]
                    if pairs:
                        lines.append(f"Row {row_number}: " + "; ".join(pairs))
                if not lines:
                    continue
                row_start = group[0][0]
                row_end = group[-1][0]
                chunks.append(
                    ParsedChunk(
                        text=f"Sheet: {sheet.title}\nColumns: {', '.join(headers)}\n" + "\n".join(lines),
                        chunk_index=len(chunks),
                        section_title=sheet.title,
                        sheet_name=sheet.title,
                        row_start=row_start,
                        row_end=row_end,
                        column_headers=tuple(headers),
                        detected_table_id=f"{sheet.title}:{row_start}-{row_end}",
                    )
                )

        if not chunks:
            raise DocumentProcessingError(f"No extractable worksheet content was found in '{filename}'.")

        return ParsedDocument(
            document_type="xlsx",
            title=title or Path(filename).stem,
            chunks=tuple(chunks),
            parser_name=PARSER_NAME,
            parser_version=PARSER_VERSION,
            sheet_names=tuple(sheet_names),
        )


def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
